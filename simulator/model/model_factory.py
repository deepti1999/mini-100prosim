import re
import abc
import pandas as pd
import sys
import os

from openpyxl import load_workbook
from ..value.value_factory import ValueFactory
from ..value.value import *
from ..value.value_collection import ValueCollection
from oemof.solph import Bus, Flow, Model, EnergySystem
from oemof.solph.components import Source, Sink, Transformer


class ModelFactory(abc.ABC):

    @property
    @abc.abstractmethod
    def value_collection(self) -> ValueCollection:
        pass

    @property
    @abc.abstractmethod
    def model(self) -> Model:
        pass


class TheModelFactory(ModelFactory):

    def __init__(self, model_name: str):
        if model_name == 'simple_model_1':
            self.__value_collection = ValueCollection(SimpleValueFactory1())
            self.__model = self.__get_simple_model_1()
        elif model_name == 'simple_model_2':
            self.__value_collection = ValueCollection(SimpleValueFactory2())
            self.__model = self.__get_simple_model_2()
        elif model_name == 'simple_model_3':
            self.__value_collection = ValueCollection(SimpleValueFactory3())
            self.__model = self.__get_simple_model_3()

    @property
    def value_collection(self) -> ValueCollection:
        return self.__value_collection

    @property
    def model(self) -> Model:
        return self.__model

    def __get_simple_model_1(self):
        """
        Simulation of a simple model to get started with Oemof:
        -------------------------------------------------------

                                              |
        Solar PV FF (150 GWh) --------------->|
        Formula: Src_PV_FF = D015 * D100      |
                                              |---------------> El. Consumer (300 GWh)
                                              |                 Value: Snk_EL_CONS
        Solar PV DF (100 GWh) --------------->|
        Value: Src_PV_DF                      |
                                              |

        Optimization: Result for Src_PV_FF = --> D015 = 2,0!

        """
        # create energy system:
        es = EnergySystem(timeindex=pd.date_range('1/1/2021', periods=1, freq='D'))

        # and add electrical bus:
        b_el = Bus(label="el_bus")
        es.add(b_el)

        #
        # --- prepare oemof entities and add them to energySystem
        #
        var_pv_ff = 'Src_PV_FF'
        src_pv_ff = self.__value_collection.value(var_pv_ff).value
        es.add(
            Source(label=var_pv_ff,
                   outputs={b_el: Flow(nominal_value=int(src_pv_ff), max=1000)})  # default: max = 1
        )
        var_pv_df = 'Src_PV_DF'
        src_pv_df = self.__value_collection.value(var_pv_df).value
        es.add(
            Source(label=var_pv_df,
                   outputs={b_el: Flow(nominal_value=int(src_pv_df), fix=1)})  # value is fixed in optimization
        )
        var_el_cons = 'Snk_EL_CONS'
        snk_el_cons = self.__value_collection.value(var_el_cons).value
        es.add(
            Sink(label=var_el_cons,
                 inputs={b_el: Flow(nominal_value=int(snk_el_cons), fix=1)})  # value is fixed in optimization
        )

        #
        # --- set up the oemof model:
        #
        return Model(energysystem=es)

    def __get_simple_model_2(self):
        """
        Simulation of a simple model to get started with Oemof:
        -------------------------------------------------------
                     b_scr1                 b1
        Src1 (100) ----|---->Tr1 (0.5)----->|
                                            |------------------------>Snk1 (300)
                                            |
        Src2 (200) -------------------------|                  b2
                                            |----Tr2 (0.5)---->|
                                                               |----->Snk2 (400)
        """
        # # container for instantiated nodes
        # noded = {}
        # # create natural gas bus
        # noded["bgas"] = solph.Bus(label="natural_gas")

        # create energy system:
        es = EnergySystem(timeindex=pd.date_range('1/1/2021', periods=1, freq='D'))

        # and add busses:
        b_src1 = Bus(label='b_src1')   # helper for src1
        es.add(b_src1)
        b1 = Bus(label="b1")
        es.add(b1)
        b2 = Bus(label="b2")
        es.add(b2)
        # --- Src1
        nam_src1 = 'Src1'
        val_src1 = self.__value_collection.value(nam_src1).value
        src1 = Source(label=nam_src1,
                      outputs={b_src1: Flow(nominal_value=int(val_src1), max=1000)})  # default: max = 1
        es.add(src1)
        # --- Tr1
        tr1 = Transformer(label="Tr1",
                          inputs={b_src1: Flow()},
                          outputs={b1: Flow()},
                          conversion_factors={b_src1: 0.5})
        es.add(tr1)
        # --- Src2
        nam_src2 = 'Src2'
        val_src2 = self.__value_collection.value(nam_src2).value
        src2 = Source(label=nam_src2,
                      outputs={b1: Flow(nominal_value=int(val_src2), fix=1)})  # value is fixed in optimization
        es.add(src2)
        # --- Snk1
        nam_snk1 = 'Snk1'
        val_snk1 = self.__value_collection.value(nam_snk1).value
        snk1 = Sink(label=nam_snk1,
                    inputs={b1: Flow(nominal_value=int(val_snk1), fix=1)})  # value is fixed in optimization
        es.add(snk1)
        # --- Snk2
        nam_snk2 = 'Snk2'
        val_snk2 = self.__value_collection.value(nam_snk2).value
        snk2 = Sink(label=nam_snk2,
                    inputs={b2: Flow(nominal_value=int(val_snk2), fix=1)})  # value is fixed in optimization
        es.add(snk2)
        # --- Tr2
        tr2 = Transformer(label="Tr2",
                          inputs={b1: Flow()},
                          outputs={b2: Flow()},
                          conversion_factors={b1: 0.5})
        es.add(tr2)

        #
        # --- set up the oemof model:
        #
        return Model(energysystem=es)

    def __get_simple_model_3(self):
        """
        Simulation of a simple model to get started with Oemof (try behaviour of transformer):
        -------------------------------------------------------
                     b_scr1                 b1
        Src1 (1000) ---|---->Tr1 (0.6)----->|------------------------>Snk1 (300) (fix)
          --> (500)              (0.4)               b2
                                   |---------------->|--------------->Snk2 (400) --> (200)
        """
        # create energy system:
        es = EnergySystem(timeindex=pd.date_range('1/1/2021', periods=1, freq='D'))

        # and add busses:
        b_src1 = Bus(label='b_src1')   # helper for src1
        es.add(b_src1)
        b1 = Bus(label="b1")
        es.add(b1)
        b2 = Bus(label="b2")
        es.add(b2)
        # --- Src1
        nam_src1 = 'Src1'
        val_src1 = self.__value_collection.value(nam_src1).value
        src1 = Source(label=nam_src1,
                      outputs={b_src1: Flow(nominal_value=int(val_src1), max=100)})
        es.add(src1)
        # --- Tr1
        tr1 = Transformer(label="Tr1",
                          inputs={b_src1: Flow()},
                          outputs={b1: Flow(), b2: Flow()},
                          conversion_factors={b1: 0.6, b2: (1-0.6)})
        es.add(tr1)
        # --- Snk1
        nam_snk1 = 'Snk1'
        val_snk1 = self.__value_collection.value(nam_snk1).value
        snk1 = Sink(label=nam_snk1,
                    inputs={b1: Flow(nominal_value=int(val_snk1), fix=1)})
        es.add(snk1)
        # --- Snk2
        nam_snk2 = 'Snk2'
        val_snk2 = self.__value_collection.value(nam_snk2).value
        snk2 = Sink(label=nam_snk2,
                    inputs={b2: Flow(nominal_value=int(val_snk2), max=100)})
        es.add(snk2)
        #
        # --- set up the oemof model:
        #
        return Model(energysystem=es)


class SimpleValueFactory3(ValueFactory):

    def value(self, vid) -> Value:
        if vid == "Src1":
            return SimpleValue(vid, 1000, Unit.GWh)
        elif vid == "Snk1":
            return SimpleValue(vid, 300, Unit.GWh)
        elif vid == "Snk2":
            return SimpleValue(vid, 400, Unit.GWh)


class SimpleValueFactory2(ValueFactory):

    def value(self, vid) -> Value:
        if vid == "Src1":
            return SimpleValue(vid, 100, Unit.GWh)
        elif vid == "Src2":
            return SimpleValue(vid, 200, Unit.GWh)
        elif vid == "Snk1":
            return SimpleValue(vid, 300, Unit.GWh)
        elif vid == "Snk2":
            return SimpleValue(vid, 400, Unit.GWh)


class SimpleValueFactory1(ValueFactory):

    def value(self, vid) -> Value:
        if vid == "Src_PV_FF":
            return FormulaValue(vid, 'D015*D100', Unit.GWh, 'D015', self)
        elif vid == "D100":
            return SimpleValue(vid, 100, Unit.GWh)
        elif vid == "D015":
            return SimpleValue(vid, 1.5, Unit.noUnit)
        elif vid == "Src_PV_DF":
            return SimpleValue(vid, 100, Unit.GWh)
        elif vid == "Snk_EL_CONS":
            return SimpleValue(vid, 300, Unit.GWh)


class ExcelModelFactory(ModelFactory, ValueFactory):

    def __init__(self, file_path, sheet_name):
        self.__wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        self.__sheet = self.__wb[sheet_name]

        first_row = next(self.__sheet.iter_rows(min_row=1, max_row=1))  # get the first row
        self.__headings = [c.value for c in first_row]  # extract the values from the cells

        self.__value_collection = ValueCollection(self)
        self.__entities = {}
        self.__model = self.__create_model()

    @property
    def value_collection(self) -> ValueCollection:
        return self.__value_collection

    @property
    def model(self) -> Model:
        return self.__model

    @property
    def entities(self):
        return self.__entities

    def value(self, vid) -> Value:
        for row_cells in self.__sheet.iter_rows(min_row=2):
            name_in_sheet = self.__get_value_from_cell(row_cells, 'Name').strip()  # Strip whitespace
            if name_in_sheet == vid:  # TODO only works for first line of transformer
                value = self.__get_value_from_cell(row_cells, 'Value')
                unit = Unit(self.__get_value_from_cell(row_cells, 'Unit'))
                # if value == '':
                #     return None
                if not re.search('[a-zA-Z]', str(value)):  # not numeric!
                    return SimpleValue(vid, value, unit)
                else:
                    free_id = self.__get_value_from_cell(row_cells, 'Free Parameter')
                    return FormulaValue(vid, value, unit, free_id, self)
        # nothing found in 'Name' column --> must be formula:
        # vname = 'vid_1'  # vid.replace('/','_').replace('-','_')  # does not work!
        # return FormulaValue(vname, vid, '', '', self)  # TODO unit missing here

    def __get_value_from_cell(self, row_cells, name):
        idx = self.__headings.index(name)
        value = row_cells[idx].value
        if value is None:
            value = ''
        elif isinstance(value, str):
            value = value.strip()  # Strip whitespace from string values
        return value

    def __create_model(self):
        es = EnergySystem(timeindex=pd.date_range('1/1/2021', periods=1, freq='D'))

        entity_name = ''
        entity_type = ''
        entity_list = []
        for row_cells in self.__sheet.iter_rows(min_row=2):
            if 'Ignore' in self.__headings and self.__get_value_from_cell(row_cells, 'Ignore') != '':
                continue

            name_xls = self.__get_value_from_cell(row_cells, 'Name')

            if name_xls != '' and name_xls != entity_name:  # new entity
                if entity_type == 'Source':
                    self.__add_source_to_model(entity_list, es)
                elif entity_type == 'Transformer':
                    self.__add_transformer_to_model(entity_list, es)
                elif entity_type == 'Sink':
                    self.__add_sink_to_model(entity_list, es)

                entity_name = name_xls
                entity_type = self.__get_value_from_cell(row_cells, 'Type')
                entity_list.clear()
                entity_list.append(row_cells)
            elif name_xls == '':  # merged cell: simply add row to list:
                entity_list.append(row_cells)

        # treat last entity:
        if entity_type == 'Source':
            self.__add_source_to_model(entity_list, es)
        elif entity_type == 'Transformer':
            self.__add_transformer_to_model(entity_list, es)
        elif entity_type == 'Sink':
            self.__add_sink_to_model(entity_list, es)

        # create the model from energy system:
        return Model(energysystem=es)

    def __add_source_to_model(self, row_list, es: EnergySystem):
        main_row = row_list[0]
        if self.__get_value_from_cell(main_row, 'Type') != 'Source':
            return

        outputs = {}
        src_name = self.__get_value_from_cell(main_row, 'Name')
        for row_cells in row_list:
            bus_name = self.__get_value_from_cell(row_cells, 'Output')
            bus = self.__get_bus(bus_name, es)
            outputs[bus] = self.__get_flow(src_name, row_cells)

        src = Source(label=src_name, outputs=outputs)
        es.add(src)
        self.__entities[src_name] = src

    def __add_transformer_to_model(self, row_list, es: EnergySystem):
        main_row = row_list[0]
        if self.__get_value_from_cell(main_row, 'Type') != 'Transformer':
            return

        inputs = {}
        outputs = {}
        conv_factors = {}
        tr_name = self.__get_value_from_cell(main_row, 'Name')
        for row_cells in row_list:
            is_input = False
            is_output = False
            bus = None
            weight = self.__get_value_from_cell(row_cells, 'Weight')
            # prepare bus:
            bus_name = self.__get_value_from_cell(row_cells, 'Input')
            if bus_name != '':
                is_input = True
                bus = self.__get_bus(bus_name, es)
            else:
                bus_name = self.__get_value_from_cell(row_cells, 'Output')
                if bus_name != '':
                    is_output = True
                    bus = self.__get_bus(bus_name, es)
            # check if value is provided
            tr_value_name = self.__get_value_from_cell(row_cells, 'Value')
            flow = self.__get_flow(tr_value_name, row_cells)
            # prepare input arguments for transformer
            if is_input and bus is not None:
                inputs[bus] = flow
            if is_output and bus is not None:
                outputs[bus] = flow
            if weight != '':
                conv_factors[bus] = self.__get_conv_factor(weight)

        tr = Transformer(label=tr_name, inputs=inputs, outputs=outputs, conversion_factors=conv_factors)
        # tr.conversion_factors = conv_factors  # TODO does not work?!
        self.__entities[tr_name] = tr
        es.add(tr)

    def __add_sink_to_model(self, row_list, es: EnergySystem):
        main_row = row_list[0]
        if self.__get_value_from_cell(main_row, 'Type') != 'Sink':
            return

        inputs = {}
        snk_name = self.__get_value_from_cell(main_row, 'Name')
        for row_cells in row_list:
            bus_name = self.__get_value_from_cell(row_cells, 'Input')
            bus = self.__get_bus(bus_name, es)
            inputs[bus] = self.__get_flow(snk_name, row_cells)

        snk = Sink(label=snk_name, inputs=inputs)
        self.__entities[snk_name] = snk
        es.add(snk)

    def __get_flow(self, value_name, row_cells) -> Flow:
        if value_name == '':
            return Flow()

        value = self.__value_collection.value(value_name).value
        free_par_name = self.__get_value_from_cell(row_cells, 'Free Parameter')

        if free_par_name == '':
            return Flow(nominal_value=int(value), fix=1)
        else:
            if value_name.endswith('_excess'):  # for 'excess' sources and sinks use high variable costs:
                return Flow(nominal_value=int(value), max=100, variable_costs=1000)
            else:
                return Flow(nominal_value=int(value), max=100)

    def __get_bus(self, bus_name, es: EnergySystem):
        if self.__entities.get(bus_name) is None:
            new_bus = Bus(label=bus_name)
            self.__entities[bus_name] = new_bus  # add new bus to internal memory
            es.add(new_bus)                      # add new bus to energy system

        return self.__entities[bus_name]

    def __get_conv_factor(self, value_or_name):
        if re.search('[a-zA-Z]', str(value_or_name)):  # not numeric!
            return self.__value_collection.value(value_or_name).value
        else:  # numeric
            return value_or_name
