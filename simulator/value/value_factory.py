import re
from .value import *
from openpyxl import load_workbook


class ValueFactory(abc.ABC):

    @abc.abstractmethod
    def value(self, vid) -> Value:
        raise NotImplementedError("Should have implemented this")


class MockupSimpleValueFactory(ValueFactory):

    def value(self, vid) -> Value:
        if vid == "d100":
            return SimpleValue(vid, 100, Unit.GWh)
        elif vid == "d10":
            return SimpleValue(vid, 10, Unit.GWh)
        elif vid == "d02":
            return SimpleValue(vid, 0.2, Unit.noUnit)


class MockupAdvancedValueFactory(ValueFactory):

    def __init__(self, vf):
        self.__valueFactory = vf

    def value(self, vid) -> Value:
        if vid == 'd30':
            return FormulaValue(vid, 'd100*d02+d10', Unit.GWh, 'd02', self)
        elif vid == 'S120*S148/100*S153/1000':
            return FormulaValue('SOL_DF_STATUS', 'S120*S148/100*S153/1000', Unit.GWh, 'S120', self)
        elif vid == 'G120*G148/100*G153/1000':
            return FormulaValue('SOL_DF_GOAL', 'G120*G148/100*G153/1000', Unit.GWh, 'G120', self)
        elif vid == 'G26*(1-G28/100)*G31/1000':
            return FormulaValue('SOL_DF_GOAL', 'G26*(1-G28/100)*G31/1000', Unit.GWh, 'G28', self)
        else:
            return self.__valueFactory.value(vid)


class XlsValueFactory(ValueFactory):

    def __init__(self, file_path, is_d_file: bool):
        self.__wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        if is_d_file:
            self.__sheet = self.__wb['1.']
            self.__column_unit = 'J'
            self.__column_status = 'K'
            self.__column_goal = 'M'
        else:
            self.__sheet = self.__wb['O_']
            self.__column_unit = 'K'
            self.__column_status = 'L'
            self.__column_goal = 'N'

    def value(self, vid) -> Value:
        if vid[0] == 'S':     # Status is demanded
            col_value = self.__column_status
        elif vid[0] == 'G':   # Goal is demanded
            col_value = self.__column_goal
        else:
            col_value = ''

        if len(col_value) > 0:
            row = vid[1:]
            value = self.__sheet[col_value+row].value
            # unit = self.__sheet[self.__column_unit+row].value
            # if unit == '%':
            #     return SimpleValue(vid, value/100, Unit.noUnit)
            # else:
            return SimpleValue(vid, value, Unit.noUnit)  # TODO units missing


# class ConfigValueFactory(ValueFactory):
#
#     def __init__(self, file_path, sheet_name):
#         self.__wb = load_workbook(filename=file_path, data_only=True, read_only=True)
#         self.__sheet = self.__wb[sheet_name]
#         first_row = next(self.__sheet.iter_rows(min_row=1, max_row=1))  # get the first row
#         self.__headings = [c.value for c in first_row]  # extract the values from the cells
#
#     def value(self, vid) -> Value:
#         for row_cells in self.__sheet.iter_rows(min_row=2):
#             if self.__get_value_from_cell(row_cells, 'Type') == 'Parameter' and \
#                self.__get_value_from_cell(row_cells, 'Name') == vid:
#                 value = self.__get_value_from_cell(row_cells, 'Value')
#                 unit = Unit(self.__get_value_from_cell(row_cells, 'Unit'))
#                 if not re.search('[a-zA-Z]', str(value)):  # not numeric!
#                     return SimpleValue(vid, value, unit)
#                 else:
#                     free_id = self.__get_value_from_cell(row_cells, 'Free Parameter')
#                     return FormulaValue(vid, value, unit, free_id, self)
#
#     def __get_value_from_cell(self, row, name):
#         idx = self.__headings.index(name)
#         value = row[idx].value
#         if value is None:
#             value = ''
#         return value

